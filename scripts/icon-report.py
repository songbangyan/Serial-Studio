#!/usr/bin/env python3
"""Icon coverage & quality report for the spec-0028 icon registry.

Three analyses over app/rcc/icons plus the command manifests and QML/C++ call sites:

  1. GAPS      - an icon is requested (a command icon renders at 32 in the palette; a
                 literal icon()/iconById() call requests a specific px) at a size no
                 available tier reaches, so the registry upscales the largest tier and it
                 looks blurry. These are the icons to draw at a larger tier.
  2. TIER HOLES- an icon has a tier above a missing one (e.g. 16 and 48 but no 24/32):
                 an inconsistent ladder that usually means a size was forgotten.
  3. RESCALES  - two tiers of one icon are pixel-identical when rendered to a common size,
                 i.e. the bigger tier is just the smaller one scaled up, adding no real
                 detail ("cheating"). Detected by rasterizing every tier to a common px
                 (cairosvg) and comparing (PIL). Needs both installed; skipped otherwise.

Outputs (next to the repo root):
  icons-to-download.csv   the GAPS list (stable schema, machine-readable)
  icon-report.html        a filterable/sortable dashboard of all three (no dependencies)
  icon-report.xlsx        the same, if openpyxl is importable (optional)

Usage:
    scripts/icon-report.py [--no-rescale] [--raster-px N]
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import subprocess
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
ICONS = ROOT / "app" / "rcc" / "icons"
CMDDIR = ROOT / "app" / "rcc" / "commands"
SRC_ROOTS = (ROOT / "app" / "qml", ROOT / "app" / "src")
TIERS = (16, 24, 32, 48)
PALETTE_PX = (
    32  # command palette grid + Start menu render every command icon at this size
)


# ------------------------------------------------------------------ inputs
def load_tiers() -> dict[tuple[str, str], set[int]]:
    tiers: dict[tuple[str, str], set[int]] = defaultdict(set)
    for path in ICONS.rglob("*.svg"):
        parts = path.relative_to(ICONS).parts
        if len(parts) == 3 and parts[1].isdigit():
            tiers[(parts[0], parts[2][:-4])].add(int(parts[1]))
    return tiers


def load_demand() -> dict[tuple[str, str], dict[int, set[str]]]:
    demand: dict[tuple[str, str], dict[int, set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )
    for mf in CMDDIR.glob("*.json"):
        try:
            data = json.loads(mf.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        for cmd in data.get("commands", []):
            for key in ("icon", "iconChecked"):
                ref = cmd.get(key, "")
                if "/" in ref:
                    cat, name = ref.split("/", 1)
                    demand[(cat, name)][PALETTE_PX].add(f"cmd:{cmd.get('id', '?')}")
    for lf in CMDDIR.glob("layouts/*.json"):
        for cat, name in re.findall(
            r'"icon"\s*:\s*"([a-z0-9]+)/([^"]+)"', lf.read_text("utf-8")
        ):
            demand[(cat, name)][PALETTE_PX].add(f"layout:{lf.name}")
    re_a = re.compile(
        r'\.icon(?:Path)?\(\s*(?:QStringLiteral\()?"([a-z0-9]+)"\)?\s*,\s*'
        r'(?:QStringLiteral\()?"([^"]+)"\)?\s*,\s*(\d+)\s*\)'
    )
    re_b = re.compile(r'\.iconById\(\s*"([a-z0-9]+)/([^"]+)"\s*,\s*(\d+)\s*\)')
    re_c = re.compile(
        r'\.iconById\([^,]*?"([a-z0-9]+)/([^"]+)"\s*:\s*"([a-z0-9]+)/([^"]+)"\s*,\s*(\d+)\s*\)'
    )
    for root in SRC_ROOTS:
        for path in root.rglob("*"):
            if path.suffix not in (".qml", ".cpp", ".h"):
                continue
            text = path.read_text(encoding="utf-8", errors="replace")
            label = path.name
            for m in re_a.finditer(text):
                demand[(m.group(1), m.group(2))][int(m.group(3))].add(label)
            for m in re_b.finditer(text):
                demand[(m.group(1), m.group(2))][int(m.group(3))].add(label)
            for m in re_c.finditer(text):
                px = int(m.group(5))
                demand[(m.group(1), m.group(2))][px].add(label)
                demand[(m.group(3), m.group(4))][px].add(label)
    return demand


# ------------------------------------------------------------------ analyses
def served(tiers: set[int], px: int) -> int | None:
    ladder = sorted(tiers)
    if not ladder:
        return None
    for tier in ladder:
        if tier >= px:
            return tier
    return ladder[-1]


def find_gaps(tiers, demand) -> list[dict]:
    rows = []
    for (cat, name), pxmap in demand.items():
        have = sorted(tiers.get((cat, name), set()))
        target = max(pxmap)
        sv = served(set(have), target)
        if sv is None:
            severity = "MISSING"
        elif sv < target:
            severity = "UPSCALE"
        elif target in have:
            continue
        elif sv >= target * 1.5:
            severity = "HEAVY_DOWNSCALE"
        else:
            continue
        rows.append(
            {
                "severity": severity,
                "category": cat,
                "name": name,
                "target_px": target,
                "download_tier": target,
                "existing_tiers": ";".join(map(str, have)) or "(none)",
                "served_tier": sv or "",
                "callsite_count": sum(len(v) for v in pxmap.values()),
                "demand_sources": ";".join(
                    f"{px}:{','.join(sorted(pxmap[px]))[:60]}" for px in sorted(pxmap)
                ),
            }
        )
    order = {"MISSING": 0, "UPSCALE": 1, "HEAVY_DOWNSCALE": 2}
    rows.sort(key=lambda r: (order[r["severity"]], r["category"], r["name"]))
    return rows


def find_holes(tiers) -> list[dict]:
    rows = []
    for (cat, name), have in tiers.items():
        present = sorted(have)
        if len(present) < 2:
            continue
        span = [t for t in TIERS if present[0] <= t <= present[-1]]
        missing = [t for t in span if t not in have]
        if missing:
            rows.append(
                {
                    "category": cat,
                    "name": name,
                    "have": ";".join(map(str, present)),
                    "missing": ";".join(map(str, missing)),
                }
            )
    rows.sort(key=lambda r: (r["category"], r["name"]))
    return rows


def rasterize(path: Path, px: int, tmp: Path):
    from PIL import Image

    png = tmp / (path.stem + f"-{path.parent.name}.png")
    out = subprocess.run(
        [
            "cairosvg",
            str(path),
            "-o",
            str(png),
            "--output-width",
            str(px),
            "--output-height",
            str(px),
        ],
        capture_output=True,
    )
    if out.returncode != 0 or not png.exists():
        return None
    img = Image.open(png).convert("RGBA")
    img.load()
    return img


def mean_diff(a, b) -> float:
    from PIL import ImageChops

    diff = ImageChops.difference(a, b)
    hist = diff.histogram()
    pixels = a.width * a.height
    total = 0
    for ch in range(len(hist) // 256):
        total += sum(i * c for i, c in enumerate(hist[ch * 256 : ch * 256 + 256]))
    return total / (pixels * (len(hist) // 256))


def find_rescales(tiers, raster_px: int) -> list[dict]:
    try:
        import PIL  # noqa: F401
    except ImportError:
        print("rescale-detection skipped: PIL not available")
        return []
    if not subprocess.run(["cairosvg", "--help"], capture_output=True).returncode == 0:
        print("rescale-detection skipped: cairosvg not available")
        return []

    import tempfile

    rows = []
    tmpdir = Path(tempfile.mkdtemp(prefix="icon-raster-"))
    for (cat, name), have in sorted(tiers.items()):
        ladder = sorted(have)
        if len(ladder) < 2:
            continue
        images = {}
        for tier in ladder:
            img = rasterize(ICONS / cat / str(tier) / f"{name}.svg", raster_px, tmpdir)
            if img is not None:
                images[tier] = img
        keys = sorted(images)
        for i in range(len(keys)):
            for j in range(i + 1, len(keys)):
                lo, hi = keys[i], keys[j]
                delta = mean_diff(images[lo], images[hi])
                if delta < 1.0:
                    rows.append(
                        {
                            "category": cat,
                            "name": name,
                            "smaller": lo,
                            "bigger": hi,
                            "mean_diff": round(delta, 3),
                            "note": f"tier {hi} is tier {lo} scaled up (no added detail)",
                        }
                    )
    rows.sort(key=lambda r: (r["mean_diff"], r["category"], r["name"]))
    return rows


# ------------------------------------------------------------------ outputs
def write_csv(gaps) -> None:
    fields = [
        "severity",
        "category",
        "name",
        "target_px",
        "download_tier",
        "existing_tiers",
        "served_tier",
        "callsite_count",
        "demand_sources",
    ]
    with open(ROOT / "icons-to-download.csv", "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(gaps)


def write_html(gaps, holes, rescales) -> None:
    def table(tid, cols, rows, sev_col=None):
        head = "".join(
            f"<th onclick=\"srt('{tid}',{i})\">{c}</th>" for i, c in enumerate(cols)
        )
        body = []
        for r in rows:
            cells = []
            for c in cols:
                v = html.escape(str(r.get(c.lower().replace(" ", "_"), "")))
                if c == sev_col:
                    v = f'<span class="pill {r[sev_col.lower()]}">{v}</span>'
                cells.append(f"<td>{v}</td>")
            body.append(f"<tr>{''.join(cells)}</tr>")
        return (
            f"<h2>{tid.capitalize()} <small>({len(rows)})</small></h2>"
            f'<input class="flt" onkeyup="flt(\'{tid}\',this.value)" placeholder="filter {tid}...">'
            f'<div class="wrap"><table id="{tid}"><thead><tr>{head}</tr></thead>'
            f'<tbody>{"".join(body)}</tbody></table></div>'
        )

    gap_cols = [
        "Severity",
        "Category",
        "Name",
        "Target Px",
        "Existing Tiers",
        "Served Tier",
        "Callsite Count",
        "Demand Sources",
    ]
    hole_cols = ["Category", "Name", "Have", "Missing"]
    resc_cols = ["Category", "Name", "Smaller", "Bigger", "Mean Diff", "Note"]
    doc = f"""<!doctype html><meta charset="utf-8"><title>Icon Report</title>
<style>
 :root{{--bg:#0e1116;--fg:#e6edf3;--mut:#8b949e;--line:#26303c;--acc:#3ddc97}}
 @media(prefers-color-scheme:light){{:root{{--bg:#fff;--fg:#1c2128;--mut:#57606a;--line:#d0d7de;--acc:#1a7f5a}}}}
 body{{background:var(--bg);color:var(--fg);font:14px/1.5 ui-sans-serif,system-ui;margin:0;padding:24px;max-width:1200px}}
 h1{{font-size:20px;margin:0 0 4px}} h2{{font-size:15px;margin:28px 0 8px}} small{{color:var(--mut)}}
 .flt{{width:100%;max-width:340px;padding:6px 10px;margin-bottom:8px;background:transparent;color:var(--fg);border:1px solid var(--line);border-radius:6px}}
 .wrap{{overflow-x:auto;border:1px solid var(--line);border-radius:8px}}
 table{{border-collapse:collapse;width:100%;font-variant-numeric:tabular-nums}}
 th,td{{text-align:left;padding:6px 10px;border-bottom:1px solid var(--line);white-space:nowrap}}
 th{{position:sticky;top:0;background:var(--bg);cursor:pointer;user-select:none;font-size:12px;color:var(--mut)}}
 td:nth-child(n+4){{font-family:ui-monospace,monospace}}
 .pill{{padding:1px 8px;border-radius:99px;font-size:12px;font-weight:600}}
 .upscale{{background:#7a1f2b;color:#ffd7dc}} .missing{{background:#7a1f2b;color:#fff}}
 .heavy_downscale{{background:#7a5a1f;color:#ffe9c7}}
 @media(prefers-color-scheme:light){{.upscale,.missing{{background:#ffdce0;color:#7a1f2b}}.heavy_downscale{{background:#fff3d6;color:#7a5a1f}}}}
</style>
<h1>Icon Report <small>&nbsp;{len(gaps)} gaps · {len(holes)} tier-holes · {len(rescales)} rescales</small></h1>
{table('gaps', gap_cols, gaps, sev_col='Severity')}
{table('holes', hole_cols, holes)}
{table('rescales', resc_cols, rescales)}
<script>
function flt(id,q){{q=q.toLowerCase();for(const r of document.querySelectorAll('#'+id+' tbody tr'))r.style.display=r.innerText.toLowerCase().includes(q)?'':'none'}}
function srt(id,c){{const t=document.getElementById(id),b=t.tBodies[0],rs=[...b.rows];const asc=t.dataset.s!=c+'a';t.dataset.s=c+(asc?'a':'d');rs.sort((x,y)=>{{const a=x.cells[c].innerText,b2=y.cells[c].innerText,n=parseFloat(a)-parseFloat(b2);return (isNaN(n)?a.localeCompare(b2):n)*(asc?1:-1)}});rs.forEach(r=>b.appendChild(r))}}
</script>"""
    (ROOT / "icon-report.html").write_text(doc, encoding="utf-8", newline="\n")


def write_xlsx(gaps, holes, rescales) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    wb = Workbook()
    sheets = [
        (
            "Gaps",
            [
                "severity",
                "category",
                "name",
                "target_px",
                "existing_tiers",
                "served_tier",
                "callsite_count",
                "demand_sources",
            ],
            gaps,
        ),
        ("Tier Holes", ["category", "name", "have", "missing"], holes),
        (
            "Rescales",
            ["category", "name", "smaller", "bigger", "mean_diff", "note"],
            rescales,
        ),
    ]
    wb.remove(wb.active)
    for title, cols, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append([c.replace("_", " ").title() for c in cols])
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"
        ws.freeze_panes = "A2"
    wb.save(ROOT / "icon-report.xlsx")
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--no-rescale", action="store_true", help="skip raster cheat-detection"
    )
    parser.add_argument(
        "--raster-px", type=int, default=256, help="common raster size for compare"
    )
    args = parser.parse_args()

    tiers = load_tiers()
    demand = load_demand()
    gaps = find_gaps(tiers, demand)
    holes = find_holes(tiers)
    rescales = [] if args.no_rescale else find_rescales(tiers, args.raster_px)

    write_csv(gaps)
    write_html(gaps, holes, rescales)
    xlsx = write_xlsx(gaps, holes, rescales)

    print(f"gaps: {len(gaps)}  tier-holes: {len(holes)}  rescales: {len(rescales)}")
    print(
        "wrote icons-to-download.csv, icon-report.html"
        + (", icon-report.xlsx" if xlsx else "")
    )
    if not xlsx:
        print("(icon-report.xlsx skipped: openpyxl not installed)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
